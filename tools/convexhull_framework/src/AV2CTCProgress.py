#!/usr/bin/env python
## Copyright (c) 2021, Alliance for Open Media. All rights reserved
##
## This source code is subject to the terms of the BSD 3-Clause Clear License and the
## Alliance for Open Media Patent License 1.0. If the BSD 3-Clause Clear License was
## not distributed with this source code in the LICENSE file, you can obtain it
## at aomedia.org/license/software-license/bsd-3-c-c/.  If the Alliance for Open Media Patent
## License 1.0 was not distributed with this source code in the PATENTS file, you
## can obtain it at aomedia.org/license/patent-license/.
##
__author__ = "maggie.sun@intel.com, ryanlei@meta.com"

import re
import os
import openpyxl
import xlsxwriter
import shutil
import Config
from Config import QPs, DnScaleRatio, CTC_ASXLSTemplate, CTC_RegularXLSTemplate, InterpolatePieces, \
    UsePCHIPInterpolation
import Utils
from Utils import ParseCSVFile, plot_rd_curve, Interpolate_Bilinear, Interpolate_PCHIP, convex_hull
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from CalcBDRate import BD_RATE
from itertools import cycle
import pandas as pd
from tabulate import tabulate

qtys = ["psnr_y", "psnr_u", "psnr_v", "overall_psnr", "ssim_y", "ms_ssim_y",
        "vmaf", "vmaf_neg", "psnr_hvs","ciede2k", "apsnr_y", "apsnr_u",
        "apsnr_v", "overall_apsnr"]

csv_paths = {
    "v1.0.0" : "F:\\AV2-CTC-new\\AV2-CTC-v1.0.0-alt-anchor-r3.0-new\\analysis\\rdresult",
    "v2.0.0" : "F:\\AV2-CTC-new\\AV2-CTC-v2.0.0-new\\analysis\\rdresult",
    "v3.0.0" : "F:\\AV2-CTC-new\\AV2-CTC-v3.0.0-new\\analysis\\rdresult",
    "v7.0.0" : "F:\\AV2-CTC-new\\AV2-CTC-v7.0.0-new\\analysis\\rdresult",
}


start_row = {
    "AI": 2,
    "AS": 2,
    "RA": 2,
    "Still": 2,
    "LD": 50,
}

formats = {
    "v1.0.0":       ['r', '-', 'o'],
    "v2.0.0":       ['g', '-', '*'],
    "v3.0.0":       ['b', '-', '^'],
    "v7.0.0":       ['r', '-', '+'],
}

AS_formats = {
    "3840x2160": ['r', '-.', 'o'],
    "2560x1440": ['g', '-.', '*'],
    "1920x1080": ['b', '-.', '^'],
    "1280x720":  ['y', '-.', '+'],
    "960x540":   ['c', '-.', 'x'],
    "640x360":   ['k', '-.', '<'],
}

anchor = "v1.0.0"
rd_curve_pdf = "rdcurve.pdf"
colors = cycle('bgrycmk')
markers = cycle('o*^+<x')

def populate_stats_files():
    stats_files = {}
    for tag in csv_paths.keys():
        stats_files[tag] = {}
        for cfg in ["AI", "LD", "RA", "Still", "AS"]:
            if cfg == "Still":
                stats_files[tag][cfg] = os.path.join(csv_paths[tag], "RDResults_aom_av2_STILL_Preset_0.csv")
            else:
                stats_files[tag][cfg] = os.path.join(csv_paths[tag], "RDResults_aom_av2_%s_Preset_0.csv"%cfg)
    return stats_files

def WriteSheet(csv_file, sht, start_row):
    csv = open(csv_file, 'rt')
    row = start_row
    for line in csv:
        if not line.startswith('TestCfg'):
            words = re.split(',', line.strip())
            col = 1
            for word in words:
                mycell = sht.cell(row=row, column=col)
                if col >= 12 and col <= 30 and word != "":
                    mycell.value = float(word)
                else:
                    mycell.value = word
                col += 1
            row += 1
    csv.close()


def FillXlsFile():
    for tag in csv_files.keys():
        if tag == anchor:
            continue
        else:
            for cfg in csv_files[anchor].keys():
                anchor_sht_name = "Anchor-%s" % cfg
                test_sht_name = "Test-%s" % cfg
                if cfg == "AS":
                    xls_template = CTC_ASXLSTemplate
                    xls_file = "CTC_AS_%s-%s.xlsm" % (anchor, tag)
                    shutil.copyfile(xls_template, xls_file)
                    anchor_sht_name = "Anchor"
                    test_sht_name = "Test"
                elif cfg == "AI":
                    xls_template = CTC_RegularXLSTemplate
                    xls_file = "CTC_Regular_%s-%s.xlsm" % (anchor, tag)
                    shutil.copyfile(xls_template, xls_file)

                wb = openpyxl.load_workbook(filename=xls_file, read_only=False, keep_vba=True)
                anchor_sht = wb[anchor_sht_name]
                anchor_csv = csv_files[anchor][cfg]
                WriteSheet(anchor_csv, anchor_sht, start_row[cfg])

                test_sht = wb[test_sht_name]
                test_csv = csv_files[tag][cfg]
                WriteSheet(test_csv, test_sht, start_row[cfg])

                wb.save(xls_file)

def DrawIndividualRDCurve(records, anchor, pdf):
    with PdfPages(pdf) as export_pdf:
        for cfg in records[anchor].keys():
            videos = records[anchor][cfg].keys()
            for video in videos:
                if cfg == "AS":
                    DnScaledRes = [(int(3840/ratio), int(2160/ratio)) for ratio in DnScaleRatio]
                    Int_RDPoints = {}
                    # draw individual rd curves
                    for tag in records.keys():
                        Int_RDPoints[tag] = []
                        record = records[tag][cfg][video]
                        plt.figure(figsize=(15, 10))
                        plt.suptitle("%s : %s: %s" % (cfg, video, tag))

                        br = {}; apsnr = {}
                        for key in record.keys():
                            res = re.split('_', key)[0]
                            if res not in br.keys():
                                br[res] = []
                                apsnr[res] = []
                            br[res].append(record[key].bitrate)
                            apsnr[res].append(record[key].overall_apsnr)

                        for res in br.keys():
                            rdpnts = [(brt, qty) for brt, qty in zip(br[res], apsnr[res])]
                            if UsePCHIPInterpolation:
                                int_rdpnts = Interpolate_PCHIP(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            else:
                                int_rdpnts = Interpolate_Bilinear(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            Int_RDPoints[tag] += int_rdpnts
                            plot_rd_curve(br[res], apsnr[res], "overall_apsnr", res, "bitrate(Kbps)",
                                          AS_formats[res][0], AS_formats[res][1], AS_formats[res][2])
                        plt.legend(loc='lower right')
                        plt.grid(True)
                        export_pdf.savefig()
                        plt.close()

                    #draw convex hull
                    plt.figure(figsize=(15, 10))
                    plt.suptitle("%s : %s: convex hull" % (cfg, video))
                    for tag in records.keys():
                        lower, upper = convex_hull(Int_RDPoints[tag])
                        br    = [h[0] for h in upper]
                        apsnr = [h[1] for h in upper]
                        plot_rd_curve(br, apsnr, "overall_apsnr(dB)", tag, "bitrate(kbps)",
                                      formats[tag][0], formats[tag][1], formats[tag][2])
                    plt.legend(loc='lower right')
                    plt.grid(True)
                    export_pdf.savefig()
                    plt.close()
                else:
                    # bit rate to quality
                    plt.figure(figsize=(15, 10))
                    plt.suptitle("%s : %s" % (cfg, video))
                    for tag in records.keys():
                        record = records[tag][cfg][video]
                        br    = [record[key].bitrate for key in record.keys()]
                        apsnr = [record[key].overall_apsnr for key in record.keys()]
                        plot_rd_curve(br, apsnr, "overall_apsnr(dB)", tag, "bitrate(kbps)",
                                      formats[tag][0], formats[tag][1], formats[tag][2])
                    plt.legend(loc='lower right')
                    plt.grid(True)
                    export_pdf.savefig()
                    plt.close()


def DrawCombinedRDCurve(records):
    pdf = "combined_rdcurve.pdf"
    with PdfPages(pdf) as export_pdf:
        for tag in csv_files.keys():
            for cfg in csv_files[tag].keys():
                videos = records[tag][cfg].keys()
                plt.figure(figsize=(30, 30))
                plt.suptitle("%s : %s" % (tag, cfg))

                for video in videos:
                    short_name = video.split('_')[0]
                    if cfg == "AS":
                        Int_RDPoints = []
                        record = records[tag][cfg][video]
                        br = {};
                        apsnr = {}
                        for key in record.keys():
                            res = re.split('_', key)[0]
                            if res not in br.keys():
                                br[res] = []
                                apsnr[res] = []
                            br[res].append(record[key].bitrate)
                            apsnr[res].append(record[key].overall_apsnr)

                        for res in br.keys():
                            rdpnts = [(brt, qty) for brt, qty in zip(br[res], apsnr[res])]
                            if UsePCHIPInterpolation:
                                int_rdpnts = Interpolate_PCHIP(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            else:
                                int_rdpnts = Interpolate_Bilinear(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            Int_RDPoints += int_rdpnts

                        # draw convex hull
                        lower, upper = convex_hull(Int_RDPoints)
                        br = [h[0] for h in upper]
                        apsnr = [h[1] for h in upper]
                        plot_rd_curve(br, apsnr, "overall_apsnr(dB)", short_name, "bitrate(kbps)",
                                      next(colors), '-', next(markers))
                    else:
                        record = records[tag][cfg][video]
                        br = [record[key].bitrate for key in record.keys()]
                        apsnr = [record[key].overall_apsnr for key in record.keys()]
                        plot_rd_curve(br, apsnr, "overall_apsnr(dB)", short_name, "bitrate(kbps)",
                                      next(colors), '-', next(markers))

                plt.legend(loc='lower right')
                plt.grid(True)
                export_pdf.savefig()
                plt.close()

def DrawCombinedRuntime(records):
    pdf = "combined_runtime.pdf"
    with PdfPages(pdf) as export_pdf:
        for tag in csv_files.keys():
            for cfg in csv_files[tag].keys():
                if cfg == "RA":
                    continue
                videos = records[tag][cfg].keys()
                plt.figure(figsize=(30, 30))
                plt.suptitle("%s : %s" % (tag, cfg))

                for video in videos:
                    short_name = video.split('_')[0]
                    if cfg == "AS":
                        record = records[tag][cfg][video]
                        br = {};
                        enc_time = {}
                        for key in record.keys():
                            res = re.split('_', key)[0]
                            if res not in br.keys():
                                br[res] = []
                                enc_time[res] = []
                            br[res].append(record[key].bitrate)
                            enc_time[res].append(record[key].enc_time)

                        for res in br.keys():
                            plot_rd_curve(br[res], enc_time[res], "enc_time(s)", short_name+'_'+res, "bitrate(kbps)",
                                          next(colors), '-', next(markers))
                    else:
                        record = records[tag][cfg][video]
                        br = [record[key].bitrate for key in record.keys()]
                        enc_time = [record[key].enc_time for key in record.keys()]
                        plot_rd_curve(br, enc_time, "enc_time(s)", short_name, "bitrate(kbps)",
                                      next(colors), '-', next(markers))

                plt.legend(loc='lower right')
                plt.grid(True)
                export_pdf.savefig()
                plt.close()

def GetQty(record, key, qty):
    if qty == 'psnr_y':
        q = record[key].psnr_y
    elif qty == 'psnr_u':
        q = record[key].psnr_u
    elif qty == 'psnr_v':
        q = record[key].psnr_v
    elif qty == 'overall_psnr':
        q = record[key].overall_psnr
    elif qty == 'ssim_y':
        q = record[key].ssim_y
    elif qty == 'ms_ssim_y':
        q = record[key].ms_ssim_y
    elif qty == 'vmaf':
        q = record[key].vmaf_y
    elif qty == 'vmaf_neg':
        q = record[key].vmaf_y_neg
    elif qty == 'psnr_hvs':
        q = record[key].psnr_hvs
    elif qty == 'ciede2k':
        q = record[key].ciede2k
    elif qty == 'apsnr_y':
        q = record[key].apsnr_y
    elif qty == 'apsnr_u':
        q = record[key].apsnr_u
    elif qty == 'apsnr_v':
        q = record[key].apsnr_v
    elif qty == 'overall_apsnr':
        q = record[key].overall_apsnr
    else:
        assert(0)
    return q

def CalcBDRate(tag, cfg, cls, video, anchor, test):
    bdrate = {}; anchor_qty = {}; test_qty = {}
    br_anchor = []; br_test = []
    for key in anchor.keys():
        br_anchor.append(anchor[key].bitrate)
    for key in test.keys():
        br_test.append(test[key].bitrate)

    for qty in qtys:
        anchor_qty[qty] = []; test_qty[qty] = []
        for key in anchor.keys():
            anchor_qty[qty].append(GetQty(anchor, key, qty))
            test_qty[qty].append(GetQty(test, key, qty))

    bdrate["tag"] = tag
    bdrate["cfg"] = cfg
    bdrate["class"] = cls
    bdrate["video"] = video

    for qty in qtys:
        (err, bd ) = BD_RATE(qty, br_anchor, anchor_qty[qty],
                                     br_test, test_qty[qty])
        if err == 0:
            bdrate[qty] = bd
        else:
            bdrate[qty] = 0.0
    return bdrate


def CalcConvexHull(record):
    br = {}; quality = {}
    cvx_hull = {}

    for key in record.keys():
        res = re.split('_', key)[0]
        if res not in br.keys():
            br[res] = []
            quality[res] = {}

        br[res].append(record[key].bitrate)
        for q in qtys:
            if q not in quality[res].keys():
                quality[res][q] = []
            quality[res][q].append(GetQty(record, key, q))

    for q in qtys:
        Int_RDPoints = []
        for res in br.keys():
            rdpnts = [(brt, qty) for brt, qty in zip(br[res], quality[res][q])]
            if UsePCHIPInterpolation:
                int_rdpnts = Interpolate_PCHIP(rdpnts, QPs['AS'][:], InterpolatePieces, True)
            else:
                int_rdpnts = Interpolate_Bilinear(rdpnts, QPs['AS'][:], InterpolatePieces, True)
            Int_RDPoints += int_rdpnts

        lower, upper = convex_hull(Int_RDPoints)
        cvx_hull[q] = upper

    return cvx_hull

def CalcASBDRate(tag, cfg, cls, video, anchor, test):
    # first produce convex hull for each quality
    convex_hull = {}
    convex_hull['anchor'] = CalcConvexHull(anchor)
    convex_hull['test'] = CalcConvexHull(test)

    # calculate bdrate for each quality metric
    bdrate = {};
    bdrate["tag"] = tag
    bdrate["cfg"] = cfg
    bdrate["class"] = cls
    bdrate["video"] = video

    for q in qtys:
        br_anchor = list(zip(*convex_hull['anchor'][q]))[0]
        br_test = list(zip(*convex_hull['test'][q]))[0]
        quality_anchor = list(zip(*convex_hull['anchor'][q]))[1]
        quality_test = list(zip(*convex_hull['test'][q]))[1]

        (err, bd ) = BD_RATE(q, br_anchor, quality_anchor, br_test, quality_test)
        if err == 0:
            bdrate[q] = bd
        else:
            bdrate[q] = 0.0
    return bdrate

def CalcFullBDRate(anchor):
    bdrate = []; seq_time = []
    for cfg in csv_files[anchor].keys():
        for video in records[anchor][cfg].keys():
            for tag in records.keys():
                record = records[tag][cfg][video]
                time = 0; instr = 0
                for key in record.keys():
                    time += record[key].enc_time
                    instr += record[key].enc_instr
                    video_class = record[key].file_class

                total_time = {}
                total_time["tag"] = tag
                total_time["cfg"] = cfg
                total_time["class"] = video_class
                total_time["video"] = video
                total_time["time"] = time
                total_time["instr"] = instr

                seq_time.append(total_time)

                if tag == anchor:
                    continue
                if cfg == "AS":
                    bd = CalcASBDRate(tag, cfg, video_class, video, records[anchor][cfg][video], records[tag][cfg][video])
                else:
                    bd = CalcBDRate(tag, cfg, video_class, video, records[anchor][cfg][video], records[tag][cfg][video])
                bdrate.append(bd)

    return (bdrate, seq_time)


def WriteSummaryXlsFile(bdrate, seq_time, seq_instr, summary):
    csv = open(summary+".csv", "wt")
    csv.write('Video,mode')
    for qty in qtys:
        csv.write(',%s'%qty)
    csv.write(",EncTime(s),EncInstr\n")

    wb = xlsxwriter.Workbook(summary + ".xlsx")
    shts = []

    for mode in csv_files.keys():
        total = {}; avg_bdrate = {}
        sht = wb.add_worksheet(mode)
        shts.append(sht)
        row = 0; col = 0
        sht.write(row, 0, 'Video')
        for qty in qtys:
            col = qtys.index(qty) + 1
            sht.write(row, col, qty)
        sht.write(row, col + 1, "EncTime(s)")
        sht.write(row, col + 2, "EncInstr")
        row = 1
        for video in bdrate.keys():
            csv.write("%s,%s"%(video, mode))
            for qty in qtys:
                if mode == anchor:
                    csv.write(',0.0')
                elif not str(bdrate[video][mode][qty]).startswith('Error'):
                    csv.write(',%f' % bdrate[video][mode][qty])
                else:
                    csv.write(',%s' % bdrate[video][mode][qty])

            csv.write(",%f,%f\n" % (seq_time[video][mode], seq_instr[video][mode]))

            sht.write(row, 0, video)
            for qty in qtys:
                if not qty in total.keys():
                    total[qty] = 0
                    avg_bdrate[qty] = 0

                if mode != anchor and not str(bdrate[video][mode][qty]).startswith('Error'):
                    total[qty] += 1
                    avg_bdrate[qty] += bdrate[video][mode][qty]

                col = qtys.index(qty) + 1
                if mode == anchor:
                    sht.write(row, col, 0.00)
                elif (mode in bdrate[video].keys()):
                    sht.write(row, col, bdrate[video][mode][qty])
            sht.write(row, col + 1, seq_time[video][mode])
            sht.write(row, col + 2, seq_instr[video][mode])
            row += 1
        row += 1
        for qty in qtys:
            if total[qty] != 0:
                avg_bdrate[qty] /= total[qty]
            else:
                avg_bdrate[qty] = 0.0

        sht.write(row, 0, "Average")
        for qty in qtys:
            col = qtys.index(qty) + 1
            sht.write(row, col, avg_bdrate[qty])
        row += 1

    wb.close()
    csv.close()

######################################
# main
######################################
if __name__ == "__main__":
    csv_files = populate_stats_files()
    records = {}
    for tag in csv_files.keys():
        records[tag] = {}
        for test_cfg in csv_files[tag].keys():
            IgnorePerf = (test_cfg == "RA")
            records[tag][test_cfg] = ParseCSVFile(csv_files[tag][test_cfg], IgnorePerf)

    FillXlsFile()
    DrawCombinedRDCurve(records)
    DrawCombinedRuntime(records)
    DrawIndividualRDCurve(records, anchor, rd_curve_pdf)

    #Calculate BDRate and collect total time
    (bdrate, seq_time) = CalcFullBDRate(anchor)

    df = pd.DataFrame(bdrate)
    average_bdrate_by_tag = df.groupby(['tag', 'cfg', 'class']).agg({
        'psnr_y': ['mean'],
        'psnr_u': ['mean'],
        'psnr_v': ['mean'],
        'overall_psnr': ['mean'],
        'ssim_y': ['mean'],
        'ms_ssim_y':['mean'],
        'vmaf':['mean'],
        'vmaf_neg':['mean'],
        'psnr_hvs':['mean'],
        'ciede2k':['mean'],
        'apsnr_y':['mean'],
        'apsnr_u':['mean'],
        'apsnr_v':['mean'],
        'overall_apsnr':['mean']
    })

    #print(average_bdrate)
    #print(tabulate(average_bdrate, headers='keys', tablefmt='psql'))

    # Write output summary csv file
    AvgBDRATEFile = "AverageBdrate-Summary-AV1-vs-AV2.csv"
    average_bdrate_by_tag.to_csv(AvgBDRATEFile, index=True)

    average_bdrate_by_video = df.groupby(['cfg', 'class', 'video', 'tag']).agg({
        'psnr_y': ['mean'],
        'psnr_u': ['mean'],
        'psnr_v': ['mean'],
        'overall_psnr': ['mean'],
        'ssim_y': ['mean'],
        'ms_ssim_y': ['mean'],
        'vmaf': ['mean'],
        'vmaf_neg': ['mean'],
        'psnr_hvs': ['mean'],
        'ciede2k': ['mean'],
        'apsnr_y': ['mean'],
        'apsnr_u': ['mean'],
        'apsnr_v': ['mean'],
        'overall_apsnr': ['mean']
    })

    PerVideoBDRATEFile = "PerVideoBdrate-Summary-AV1-vs-AV2.csv"
    average_bdrate_by_video.to_csv(PerVideoBDRATEFile, index=True)